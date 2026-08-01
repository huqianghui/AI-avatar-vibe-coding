"""CRM Excel import service (Phase 33, PERS-01, D-01..D-04)."""

import io
from dataclasses import dataclass, field

import openpyxl
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import get_settings
from app.models.user import User
from app.models.user_crm_context import UserCrmContext
from app.services.personalization_sanitizer import sanitize_field, sanitize_free_text_with_pii

EXPECTED_HEADERS = ["user_email", "customer_name", "company", "role", "crm_notes", "contact_person"]


class CrmHeaderValidationError(ValueError):
    """Raised when the uploaded workbook's header row doesn't exactly match EXPECTED_HEADERS."""


@dataclass
class CrmImportResult:
    success_count: int = 0
    skipped: list[dict] = field(default_factory=list)
    unmatched: list[dict] = field(default_factory=list)


async def parse_and_import_crm_excel(db: AsyncSession, file_bytes: bytes) -> CrmImportResult:
    settings = get_settings()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    header_values = [str(h).strip() if h is not None else "" for h in (header or [])]
    if header_values != EXPECTED_HEADERS:
        raise CrmHeaderValidationError(f"Expected header {EXPECTED_HEADERS}, got {header_values}")

    result = CrmImportResult()
    for row_num, row in enumerate(rows_iter, start=2):
        if row is None or all(cell is None for cell in row):
            continue
        values = list(row) + [None] * (len(EXPECTED_HEADERS) - len(row))
        email = str(values[0]).strip() if values[0] else ""
        if not email:
            result.skipped.append({"row": row_num, "reason": "missing user_email"})
            continue

        user_result = await db.execute(select(User).where(func.lower(User.email) == email.lower()))
        user = user_result.scalar_one_or_none()
        if user is None:
            result.unmatched.append({"row": row_num, "email": email, "reason": "email not found"})
            continue

        existing_result = await db.execute(
            select(UserCrmContext).where(UserCrmContext.user_id == user.id)
        )
        crm_context = existing_result.scalar_one_or_none() or UserCrmContext(user_id=user.id)
        crm_context.customer_name = sanitize_field(values[1], settings.crm_field_max_length)
        crm_context.company = sanitize_field(values[2], settings.crm_field_max_length)
        crm_context.role = sanitize_field(values[3], settings.crm_field_max_length)
        crm_context.crm_notes = sanitize_free_text_with_pii(
            values[4], settings.crm_notes_max_length
        )
        crm_context.contact_person = sanitize_field(values[5], settings.crm_field_max_length)
        db.add(crm_context)
        result.success_count += 1

    await db.commit()
    return result


def generate_crm_template_workbook() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(EXPECTED_HEADERS)
    ws.append(["user@example.com", "张三", "示例医院", "主任医师", "示例备注", "李四"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
