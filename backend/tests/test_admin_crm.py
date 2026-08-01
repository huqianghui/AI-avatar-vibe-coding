"""Integration tests for admin CRM Excel import API endpoints (Phase 33, PERS-01)."""

import io

import openpyxl

from app.config import get_settings
from app.models.user import User
from app.services.auth import create_access_token, get_password_hash
from app.services.crm_import_service import EXPECTED_HEADERS
from tests.conftest import TestSessionLocal


async def _create_admin_and_token() -> tuple[str, str]:
    """Create an admin user and return (user_id, bearer_token)."""
    async with TestSessionLocal() as session:
        user = User(
            username="admin_crm",
            email="admin_crm@test.com",
            hashed_password=get_password_hash("admin123"),
            full_name="Admin CRM",
            role="admin",
        )
        session.add(user)
        await session.commit()
        await session.refresh(user)
        token = create_access_token(data={"sub": user.id})
        return user.id, token


async def _create_user_and_token() -> tuple[str, str]:
    """Create a regular (non-admin) user and return (user_id, bearer_token)."""
    async with TestSessionLocal() as session:
        user = User(
            username="user_crm",
            email="user_crm@test.com",
            hashed_password=get_password_hash("pass123"),
            full_name="Regular User",
            role="user",
        )
        session.add(user)
        await session.commit()
        await session.refresh(user)
        token = create_access_token(data={"sub": user.id})
        return user.id, token


async def _create_target_user() -> User:
    """Create a plain user whose email will match a CRM upload row."""
    async with TestSessionLocal() as session:
        user = User(
            username="target_user",
            email="target@example.com",
            hashed_password=get_password_hash("pass123"),
            full_name="Target User",
            role="user",
        )
        session.add(user)
        await session.commit()
        await session.refresh(user)
        return user


def _build_workbook(rows: list[list]) -> bytes:
    """rows[0] is the header row, remaining rows are data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestUploadCrmExcel:
    """Tests for POST /api/v1/admin/crm/upload."""

    async def test_valid_upload_returns_200_with_counts(self, client):
        await _create_target_user()
        _, token = await _create_admin_and_token()
        file_bytes = _build_workbook(
            [
                EXPECTED_HEADERS,
                ["target@example.com", "张三", "示例医院", "主任医师", "备注", "李四"],
            ]
        )

        response = await client.post(
            "/api/v1/admin/crm/upload",
            files={
                "file": (
                    "crm.xlsx",
                    io.BytesIO(file_bytes),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 200
        data = response.json()
        assert data == {"success_count": 1, "skipped": [], "unmatched": []}

    async def test_header_mismatch_returns_422_and_no_log_created(self, client):
        _, token = await _create_admin_and_token()
        file_bytes = _build_workbook(
            [
                ["user_email", "customer_name", "wrong_column", "role", "crm_notes", "contact"],
                ["target@example.com", "张三", "示例医院", "主任医师", "备注", "李四"],
            ]
        )

        response = await client.post(
            "/api/v1/admin/crm/upload",
            files={
                "file": (
                    "crm.xlsx",
                    io.BytesIO(file_bytes),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 422
        data = response.json()
        assert data["code"] == "VALIDATION_ERROR"

        last_import_response = await client.get(
            "/api/v1/admin/crm/last-import",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert last_import_response.json() is None

    async def test_oversize_file_returns_422_before_parsing(self, client, monkeypatch):
        _, token = await _create_admin_and_token()
        settings = get_settings()
        monkeypatch.setattr(settings, "crm_max_file_size_bytes", 10)

        file_bytes = _build_workbook(
            [
                EXPECTED_HEADERS,
                ["target@example.com", "张三", "示例医院", "主任医师", "备注", "李四"],
            ]
        )
        response = await client.post(
            "/api/v1/admin/crm/upload",
            files={
                "file": (
                    "crm.xlsx",
                    io.BytesIO(file_bytes),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 422

    async def test_bad_extension_returns_422(self, client):
        _, token = await _create_admin_and_token()
        response = await client.post(
            "/api/v1/admin/crm/upload",
            files={"file": ("crm.csv", io.BytesIO(b"user_email,foo\n"), "text/csv")},
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 422

    async def test_non_admin_gets_403(self, client):
        _, token = await _create_user_and_token()
        file_bytes = _build_workbook([EXPECTED_HEADERS])
        response = await client.post(
            "/api/v1/admin/crm/upload",
            files={
                "file": (
                    "crm.xlsx",
                    io.BytesIO(file_bytes),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 403

    async def test_no_auth_returns_401(self, client):
        file_bytes = _build_workbook([EXPECTED_HEADERS])
        response = await client.post(
            "/api/v1/admin/crm/upload",
            files={
                "file": (
                    "crm.xlsx",
                    io.BytesIO(file_bytes),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert response.status_code == 401


class TestDownloadCrmTemplate:
    """Tests for GET /api/v1/admin/crm/template."""

    async def test_download_template_as_admin_returns_valid_xlsx(self, client):
        _, token = await _create_admin_and_token()

        response = await client.get(
            "/api/v1/admin/crm/template",
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 200
        assert "attachment" in response.headers["content-disposition"]
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        header_row = next(ws.iter_rows(values_only=True))
        assert list(header_row) == EXPECTED_HEADERS

    async def test_download_template_no_auth_returns_401(self, client):
        response = await client.get("/api/v1/admin/crm/template")
        assert response.status_code == 401


class TestGetLastCrmImport:
    """Tests for GET /api/v1/admin/crm/last-import."""

    async def test_returns_null_before_any_upload(self, client):
        _, token = await _create_admin_and_token()

        response = await client.get(
            "/api/v1/admin/crm/last-import",
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 200
        assert response.json() is None

    async def test_returns_result_after_upload(self, client):
        await _create_target_user()
        _, token = await _create_admin_and_token()
        file_bytes = _build_workbook(
            [
                EXPECTED_HEADERS,
                ["target@example.com", "张三", "示例医院", "主任医师", "备注", "李四"],
            ]
        )
        await client.post(
            "/api/v1/admin/crm/upload",
            files={
                "file": (
                    "crm_upload.xlsx",
                    io.BytesIO(file_bytes),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"Authorization": f"Bearer {token}"},
        )

        response = await client.get(
            "/api/v1/admin/crm/last-import",
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 200
        data = response.json()
        assert data["filename"] == "crm_upload.xlsx"
        assert data["success_count"] == 1
