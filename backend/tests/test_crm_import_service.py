"""Tests for CRM Excel import service (Phase 33, PERS-01, D-01..D-04)."""

import io
import json

import openpyxl
import pytest
from sqlalchemy import select

from app.models.user import User
from app.models.user_crm_context import UserCrmContext
from app.schemas.crm_import import CrmImportLogOut
from app.services.crm_import_service import (
    EXPECTED_HEADERS,
    CrmHeaderValidationError,
    CrmImportResult,
    generate_crm_template_workbook,
    get_last_import_log,
    parse_and_import_crm_excel,
    record_import_log,
)


def _build_workbook(rows: list[list]) -> bytes:
    """rows[0] is the header row, remaining rows are data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _seed_user(db_session, email: str, username: str) -> User:
    user = User(
        username=username,
        email=email,
        hashed_password="not-a-real-hash",
    )
    db_session.add(user)
    await db_session.commit()
    await db_session.refresh(user)
    return user


class TestParseAndImportCrmExcel:
    @pytest.mark.asyncio
    async def test_valid_header_and_matched_row_upserts_one_context(self, db_session):
        user = await _seed_user(db_session, "alice@example.com", "alice")
        file_bytes = _build_workbook(
            [
                EXPECTED_HEADERS,
                ["Alice@Example.com", "张三", "示例医院", "主任医师", "普通备注", "李四"],
            ]
        )

        result = await parse_and_import_crm_excel(db_session, file_bytes)

        assert result.success_count == 1
        assert result.skipped == []
        assert result.unmatched == []

        row = (
            await db_session.execute(
                select(UserCrmContext).where(UserCrmContext.user_id == user.id)
            )
        ).scalar_one()
        assert row.customer_name == "张三"
        assert row.company == "示例医院"

    @pytest.mark.asyncio
    async def test_wrong_header_raises_before_persisting_any_row(self, db_session):
        await _seed_user(db_session, "bob@example.com", "bob")
        file_bytes = _build_workbook(
            [
                [
                    "user_email",
                    "customer_name",
                    "wrong_column",
                    "role",
                    "crm_notes",
                    "contact_person",
                ],
                ["bob@example.com", "Bob", "Acme", "Manager", "notes", "Carol"],
            ]
        )

        with pytest.raises(CrmHeaderValidationError):
            await parse_and_import_crm_excel(db_session, file_bytes)

        count = (await db_session.execute(select(UserCrmContext))).scalars().all()
        assert count == []

    @pytest.mark.asyncio
    async def test_missing_email_row_is_skipped_not_persisted_and_does_not_raise(self, db_session):
        file_bytes = _build_workbook(
            [
                EXPECTED_HEADERS,
                ["", "No Email", "Acme", "Manager", "notes", "Carol"],
            ]
        )

        result = await parse_and_import_crm_excel(db_session, file_bytes)

        assert result.success_count == 0
        assert len(result.skipped) == 1
        assert result.skipped[0]["row"] == 2
        assert "reason" in result.skipped[0]
        assert result.unmatched == []

    @pytest.mark.asyncio
    async def test_unmatched_email_row_reported_not_persisted(self, db_session):
        file_bytes = _build_workbook(
            [
                EXPECTED_HEADERS,
                ["ghost@example.com", "Ghost", "Nowhere", "Manager", "notes", "Carol"],
            ]
        )

        result = await parse_and_import_crm_excel(db_session, file_bytes)

        assert result.success_count == 0
        assert result.skipped == []
        assert len(result.unmatched) == 1
        assert result.unmatched[0]["row"] == 2
        assert result.unmatched[0]["email"] == "ghost@example.com"
        assert "reason" in result.unmatched[0]

    @pytest.mark.asyncio
    async def test_crm_notes_with_phone_number_persisted_redacted(self, db_session):
        await _seed_user(db_session, "dana@example.com", "dana")
        file_bytes = _build_workbook(
            [
                EXPECTED_HEADERS,
                [
                    "dana@example.com",
                    "Dana",
                    "Acme",
                    "Manager",
                    "联系电话13812345678",
                    "Carol",
                ],
            ]
        )

        result = await parse_and_import_crm_excel(db_session, file_bytes)

        assert result.success_count == 1
        row = (await db_session.execute(select(UserCrmContext))).scalar_one()
        assert "[PHONE_REDACTED]" in row.crm_notes
        assert "13812345678" not in row.crm_notes

    @pytest.mark.asyncio
    async def test_reupload_for_same_user_updates_existing_row_not_duplicate(self, db_session):
        user = await _seed_user(db_session, "erin@example.com", "erin")

        first_bytes = _build_workbook(
            [
                EXPECTED_HEADERS,
                ["erin@example.com", "Erin V1", "Acme", "Manager", "notes v1", "Carol"],
            ]
        )
        await parse_and_import_crm_excel(db_session, first_bytes)

        first_row = (
            await db_session.execute(
                select(UserCrmContext).where(UserCrmContext.user_id == user.id)
            )
        ).scalar_one()
        first_id = first_row.id

        second_bytes = _build_workbook(
            [
                EXPECTED_HEADERS,
                ["erin@example.com", "Erin V2", "Acme", "Manager", "notes v2", "Carol"],
            ]
        )
        result = await parse_and_import_crm_excel(db_session, second_bytes)

        assert result.success_count == 1
        rows = (
            (
                await db_session.execute(
                    select(UserCrmContext).where(UserCrmContext.user_id == user.id)
                )
            )
            .scalars()
            .all()
        )
        assert len(rows) == 1
        assert rows[0].id == first_id
        assert rows[0].customer_name == "Erin V2"


class TestGenerateCrmTemplateWorkbook:
    def test_returns_bytes_with_expected_header_row(self):
        file_bytes = generate_crm_template_workbook()

        assert isinstance(file_bytes, bytes)
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        header_row = next(ws.iter_rows(values_only=True))
        assert list(header_row) == EXPECTED_HEADERS


class TestRecordAndGetLastImportLog:
    """Tests for CrmImportLog persistence (Phase 33, PERS-01, D-12)."""

    @pytest.mark.asyncio
    async def test_record_import_log_persists_row_and_returns_it(self, db_session):
        result = CrmImportResult(
            success_count=2,
            skipped=[{"row": 3, "reason": "missing user_email"}],
            unmatched=[{"row": 5, "email": "x@x.com", "reason": "email not found"}],
        )

        log = await record_import_log(db_session, "crm.xlsx", result, "admin-1")

        assert log.id is not None
        assert log.filename == "crm.xlsx"
        assert log.success_count == 2
        assert log.imported_by == "admin-1"
        assert json.loads(log.skipped) == result.skipped
        assert json.loads(log.unmatched) == result.unmatched

    @pytest.mark.asyncio
    async def test_get_last_import_log_returns_none_when_no_rows(self, db_session):
        assert await get_last_import_log(db_session) is None

    @pytest.mark.asyncio
    async def test_get_last_import_log_returns_most_recent_by_created_at(self, db_session):
        await record_import_log(db_session, "first.xlsx", CrmImportResult(), "admin-1")
        second = await record_import_log(db_session, "second.xlsx", CrmImportResult(), "admin-1")

        last = await get_last_import_log(db_session)

        assert last is not None
        assert last.id == second.id
        assert last.filename == "second.xlsx"

    @pytest.mark.asyncio
    async def test_crm_import_log_out_from_log_json_decodes_skipped_and_unmatched(self, db_session):
        result = CrmImportResult(
            success_count=2,
            skipped=[{"row": 3, "reason": "missing user_email"}],
            unmatched=[{"row": 5, "email": "x@x.com", "reason": "email not found"}],
        )
        log = await record_import_log(db_session, "crm.xlsx", result, "admin-1")

        out = CrmImportLogOut.from_log(log)

        assert out.skipped == result.skipped
        assert out.unmatched == result.unmatched
        assert isinstance(out.skipped[0], dict)
        assert isinstance(out.unmatched[0], dict)
